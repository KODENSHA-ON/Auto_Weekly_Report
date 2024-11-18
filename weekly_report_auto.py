import os
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.worksheet.copier import WorksheetCopy
from datetime import datetime, timedelta
from oauth2client.service_account import ServiceAccountCredentials
import gspread

def get_japanese_weekday(date):
    weekdays = ['月', '火', '水', '木', '金', '土', '日']
    return weekdays[date.weekday()]

# Googleスプレッドシート認証およびサービス作成
def get_google_sheet(sheet_name, json_keyfile_path):
    scopes = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    credentials = ServiceAccountCredentials.from_json_keyfile_name(json_keyfile_path, scopes)
    gc = gspread.authorize(credentials)
    return gc.open(sheet_name).sheet1

# ファイル選択関数
def select_file():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(title='更新する週報を選択してください。', filetypes=[('Excel files', '*.xlsx')])
    return file_path

# エクセルファイル選択
file_path = select_file()
directory, original_filename = os.path.split(file_path)

# エクセルファイルをロード
workbook = load_workbook(file_path)

# 最後のシートを選択
last_sheet = workbook.worksheets[-1]

# 新しいシート名を生成（例：11.01~11.07）
last_sheet_name = last_sheet.title
start_date_str, end_date_str = last_sheet_name.split('~')
start_date = datetime.strptime(start_date_str, '%m.%d') + timedelta(days=7)
end_date = datetime.strptime(end_date_str, '%m.%d') + timedelta(days=7)
new_sheet_name = start_date.strftime('%m.%d') + '~' + end_date.strftime('%m.%d')

# シートをコピー
new_sheet = workbook.create_sheet(title=new_sheet_name)
WorksheetCopy(last_sheet, new_sheet).copy_worksheet()

# 日付を変更
start_date = datetime(2024, start_date.month, start_date.day)
end_date = datetime(2024, end_date.month, end_date.day)
new_sheet['B2'] = f"【週報】：2024年　{start_date.month}月　{start_date.day}日　～2024年　{end_date.month}月　{end_date.day}日"

# 作成日を変更
today = datetime.now()
new_sheet['H2'] = f"作成日：　{today.year}年　{today.month}月　{today.day}日"

# Googleスプレッドシートをロード
script_dir = os.path.dirname(__file__)  # スクリプトの現在のパス
json_keyfile_name = 'ace-shine-440901-t1-ed87bd18eddb.json'
json_keyfile_path = os.path.join(script_dir, json_keyfile_name)
sheet_name = '開発２課案件リスト'
sheet = get_google_sheet(sheet_name, json_keyfile_path)

# スプレッドシートのデータをフィルタリング
rows = sheet.get_all_values()[1:]  # ヘッダーを除くすべてのデータ行
relevant_rows = []
for row in rows:
    try:
        if row[6] and start_date <= datetime.strptime(row[1], '%Y/%m/%d') <= end_date and row[9]:
            relevant_rows.append(row)
    except ValueError:  # 日付形式が一致しない場合
        continue

# 日報_温.xlsxをロード
daily_report_path = r'\\kaihatuserver\Kaihatu\IT開発プロジェクト\日報\温\日報_温.xlsx'
daily_report_wb = load_workbook(daily_report_path)
daily_report_sheet = daily_report_wb.worksheets[-1]

# 日報_温.xlsxのデータをフィルタリング
daily_report_rows = daily_report_sheet.iter_rows(min_row=3, values_only=True)
daily_report_data = []
for row in daily_report_rows:
    try:
        if row[0]:
            date_str = str(row[0])
            try:
                # 'YYYY-MM-DD HH:MM:SS'
                parsed_date = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                try:
                    # 'YYYY/MM/DD' 
                    parsed_date = datetime.strptime(date_str, '%Y/%m/%d')
                except ValueError:
                    # 
                    print(f"サポートされない形式の日付: {date_str}")
                    continue
            
            if start_date <= parsed_date <= end_date:
                daily_report_data.append(row)
    except Exception as e:
        print(f"日報_温.xlsxで問題が発生しました: {str(e)}")
        print(f"不明な行: {row}")
        continue

Found_Co01104000 = False

# エクセルシートを更新
for row in new_sheet.iter_rows():
    if row[8].value == '完':
        for i in range(5, 11):
            row[i].value = None
    
    if row[1].value == 'CO01104000' and not Found_Co01104000:
        Found_Co01104000 = True
        # Fセルを更新
        start_japanese_weekday = get_japanese_weekday(start_date)
        end_japanese_weekday = get_japanese_weekday(end_date)
        row[5].value = f"● {start_date.strftime('%m/%d')}({start_japanese_weekday})-{end_date.strftime('%m/%d')}({end_japanese_weekday})：\n"
        row[5].value += "、".join([r[2] for r in relevant_rows])

        total_hours = 0
        for r in relevant_rows:
            try:
                hours = float(r[9])
                total_hours += hours
            except (ValueError, TypeError):
                total_hours += 0
        row[13].value = total_hours
    else:
        # 日報_温.xlsxのデータを反映 
        matching_daily_report_rows = [r for r in daily_report_data if r[1] == row[1].value] 
        if matching_daily_report_rows: 
            existing_value = row[5].value or ""
            lines = existing_value.split('\n')
            if lines: 
                lines.pop(0) 
            remaining_value = '\n'.join(lines)
            
            start_japanese_weekday = get_japanese_weekday(start_date) 
            end_japanese_weekday = get_japanese_weekday(end_date) 
            new_first_line = f"● {start_date.strftime('%m/%d')}({start_japanese_weekday})-{end_date.strftime('%m/%d')}({end_japanese_weekday})：\n" 
            row[5].value = new_first_line + remaining_value
            def parse_date(date_str):
                try:
                    # 'YYYY-MM-DD HH:MM:SS'
                    return datetime.strptime(str(date_str), '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    try:
                        # 'YYYY/MM/DD' 
                        return datetime.strptime(str(date_str), '%Y/%m/%d')
                    except ValueError:
                        print(f"サポートされない形式の日付: {date_str}")
                        return None

            report_str = "、".join([f"{r[3]}" for r in matching_daily_report_rows if parse_date(r[0])]) 
            row[5].value += f"、{report_str}" 
            if row[13].value is None: 
                row[13].value = 0.0
            row[13].value += sum(float(r[6]) for r in matching_daily_report_rows)

# ファイル名を変更
if "_" in original_filename:
    parts = original_filename.split('_')
    if len(parts) > 2 and len(parts[-2]) == 8 and parts[-2].isdigit():
        parts[-2] = end_date.strftime('%Y%m%d')
        new_filename = '_'.join(parts)
    else:
        new_filename = f"{'_'.join(parts[:-1])}_{end_date.strftime('%Y%m%d')}_{parts[-1]}"
else:
    new_filename = f"{original_filename}_{end_date.strftime('%Y%m%d')}.xlsx"

# 保存パスを設定
save_path = os.path.join(directory, new_filename)

# エクセルファイルを保存
workbook.save(save_path)

print(f"ファイルが正常に保存されました: {save_path}")
